#!/usr/bin/env python3
from openpyxl import load_workbook
from titlecase import titlecase
from json import dumps
import sys
import re

descriptors = [
    'acid',
    'air',
    'chaotic',
    'cold',
    'curse',
    'darkness',
    'death',
    'disease',
    'earth',
    'electricity',
    'emotion',
    'evil',
    'fear',
    'fire',
    'force',
    'good',
    'language_dependent',
    'lawful',
    'light',
    'mind_affecting',
    'pain',
    'poison',
    'shadow',
    'sonic',
    'water',
    'ruse',
    'draconic',
    'meditative'
]

classes = [
    "sor",
    "wiz",
    "cleric",
    "druid",
    "ranger",
    "bard",
    "paladin",
    "alchemist",
    "summoner",
    "witch",
    "inquisitor",
    "oracle",
    "antipaladin",
    "magus",
    "adept",
    "bloodrager",
    "shaman",
    "psychic",
    "medium",
    "mesmerist",
    "occultist",
    "spiritualist",
    "skald",
    "investigator",
    "hunter",
    "summoner_unchained"
]

timecategory = re.compile(r"…|\/.+")
one = re.compile(r"\b(one)\b")
hitdieless = re.compile(r"(\d+)\s+HD\s+or\s+less")
hitdiemore = re.compile(r"(\d+)\s+HD\s+or\s+more")
def shorten(val, aggressive = True):
    mapping = [
        ('caster level', 'level'), ('levels', 'cl'), ('level', 'cl'),  ('ft.', 'ft'), ('rounds', 'rd'), ('round', 'rd'),
        ('min.', 'min'), ('minutes', 'min'), ('minute', 'min'), ('hours', 'h'), ('hour', 'h'), ('maximum', 'max'), ('minimum', 'min'),
        ('instantaneous', 'instant'), ('Concentration', 'conc.'), ('concentration', 'conc.'),
        ('two', '2'), ('three', '3'),
        ('creature or creatures', 'creature(s)')
    ]

    for k, v in mapping:
        val = val.replace(k, v)

    val = one.sub('1', val)
    val = hitdiemore.sub('≥&#8201;\g<1>&#8201;HD', val)
    val = hitdieless.sub('≤&#8201;\g<1>&#8201;HD', val)

    if aggressive == False:
        return val

    mapping = [
        (' cl', 'cl'), (' ft', 'ft'), (' rd', 'rd'), (' min', 'min'), 
        ('unlimited', '∞'), ('permanent', '∞'), (' per ', '/'), (' (D)', '; D'),
        ('until discharged', 'discharge'), (' or discharged', '/discharge' ), ('or until', 'or'),
        (' standard', ' std'), (' action', ''),
        ('; see text', '…'), (', see text', '…'), ('; see below', '…'), (', see below', '…'), (' (see below)', '…'), (' (see text)', '…'), ('see text', '…'), ('see below', '…'), ('see Text', '…'),
        (' (object)', '; obj.'), (' (harmless)', '; safe'), (' (harmless, object)', '; safe, obj.'),
        (' (if interacted with)', ', interact'),
        ('none', 'no'), ('None', 'no' ), ('No', 'no'),
        ('Fortitude', 'Fort.'), ('Reflex', 'Ref.'), ('Fort. negates', '<del>Fort.</del>'), ('Will negates', '<del>Will</del>'), ('Ref. negates', '<del>Ref.</del>'),
        ('Fort. partial', '<i>Fort.</i>'), ('Will partial', '<i>Will</i>'), ('Ref. partial', '<i>Ref.</i>'),
        ('.…', '…'), (' and …', '…'), (' + ', '+')
    ]

    for k, v in mapping:
        val = val.replace(k, v)
    return val

def shorten_casting_time(val):
    val = shorten(val, True).strip()
    mapping = [
        ('at least ', ''),
        ('full-rd', 'full'), ('fullrd', 'full'),
        ('Casting time ', ''), (', special ', ''),
        (' or more', ''), (' or ', ''),
        ('stdimmediate', 'std'),
        ('standard', 'std'), ('rd', ' round'),
        ('min', ' min'), (' of target', ''), (' created', ''),
        (', plus length of memory to be altered', '…')
    ]

    if val[0] != "…" and val[0] != " " and val[0].isdigit() == False:
        val = "1 " + val

    # wtf?
    mapping.append(('1 1', '1'))

    for k, v in mapping:
        val = val.replace(k, v)
    return val

def conditions(val):
    mapping = [
        ('Bleed', '<strong>Bleed</strong>',), ('bleed', '<strong>bleed</strong>'), ('Blinded', '<strong>Blinded</strong>',), ('blinded', '<strong>blinded</strong>'), ('Broken', '<strong>Broken</strong>',), ('broken', '<strong>broken</strong>'),
        ('Confused', '<strong>Confused</strong>',), ('confused', '<strong>confused</strong>'), ('Cowering', '<strong>Cowering</strong>',), ('cowering', '<strong>cowering</strong>'), ('Dazed', '<strong>Dazed</strong>',), ('dazed', '<strong>dazed</strong>'), ('Dazzled', '<strong>Dazzled</strong>',), ('dazzled', '<strong>dazzled</strong>'),
        ('Dead', '<strong>Dead</strong>',), 
        ('undead', 'RECOVER'), ('dead', '<strong>dead</strong>'), ('RECOVER', 'undead'),
        ('Deafened', '<strong>Deafened</strong>',), ('deafened', '<strong>deafened</strong>'), ('Disabled', '<strong>Disabled</strong>',), ('disabled', '<strong>disabled</strong>'), ('Dying', '<strong>Dying</strong>',), ('dying', '<strong>dying</strong>'),
        ('Energy Drained', '<strong>Energy Drained</strong>',), ('energy drained', '<strong>energy drained</strong>'), ('Entangled', '<strong>Entangled</strong>',), ('entangled', '<strong>entangled</strong>'), ('Exhausted', '<strong>Exhausted</strong>',), ('exhausted', '<strong>exhausted</strong>'),
        ('Fascinated', '<strong>Fascinated</strong>',), ('fascinated', '<strong>fascinated</strong>'), ('Fatigued', '<strong>Fatigued</strong>',), ('fatigued', '<strong>fatigued</strong>'),
        ('Flat-Footed', '<strong>Flat-Footed</strong>',), ('flat-footed', '<strong>flat-footed</strong>'), ('flat footed', '<strong>flat footed</strong>'), ('Frightened', '<strong>Frightened</strong>',), ('frightened', '<strong>frightened</strong>'), ('Grappled', '<strong>Grappled</strong>',), ('grappled', '<strong>grappled</strong>'),
        ('Helpless', '<strong>Helpless</strong>',), ('helpless', '<strong>helpless</strong>'), ('Incorporeal', '<strong>Incorporeal</strong>',), ('incorporeal', '<strong>incorporeal</strong>'), ('Invisible', '<strong>Invisible</strong>',), ('invisible', '<strong>invisible</strong>'), ('Nauseated', '<strong>Nauseated</strong>',), ('nauseated', '<strong>nauseated</strong>'),
        ('Panicked', '<strong>Panicked</strong>',), ('panicked', '<strong>panicked</strong>'), ('Paralyzed', '<strong>Paralyzed</strong>',), ('paralyzed', '<strong>paralyzed</strong>'), ('Petrified', '<strong>Petrified</strong>',), ('petrified', '<strong>petrified</strong>'), ('Pinned', '<strong>Pinned</strong>',), ('pinned', '<strong>pinned</strong>'),
        ('Prone', '<strong>Prone</strong>',), ('prone', '<strong>prone</strong>'), ('Shaken', '<strong>Shaken</strong>',), ('shaken', '<strong>shaken</strong>'), ('Sickened', '<strong>Sickened</strong>',), ('sickened', '<strong>sickened</strong>'), ('Sinking', '<strong>Sinking</strong>',), ('sinking', '<strong>sinking</strong>'),
        ('Stable', '<strong>Stable</strong>',), ('stable', '<strong>stable</strong>'), ('Staggered', '<strong>Staggered</strong>',), ('staggered', '<strong>staggered</strong>'), ('Stunned', '<strong>Stunned</strong>',), ('stunned', '<strong>stunned</strong>'), ('Unconscious', '<strong>Unconscious</strong>',), ('unconscious', '<strong>unconscious</strong>')
    ]
    for k, v in mapping:
        val = val.replace(k, v)
    return val

def get_time_category(val):
    val = timecategory.sub('', val)
    if len(val) > 0 and val[0] != "1":
        return "Long"
    if "swift" in val:
        return "Swift"
    if "immediate" in val:
        return "Immediate"
    if "std" in val:
        return "Standard"
    if "full" in val:
        return "Full-round"
    if "round" in val:
        return "Round"
    return "Long"


def is_time_special(val):
    return '/' in val or '…' in val

def read_spells(file):
    wb = load_workbook(filename = file, read_only=True)
    spellSheet = wb.sheetnames[0]
    r = wb[spellSheet]
    rows = r.rows
    first_row = [cell.value for cell in next(rows)]
    cards = {}
    materials = re.compile(r"([VSMDF\/]+)\s?\((.*?)\)")
    details = re.compile(r"(\s+\(.*?\)|\s+)")
    dicedc = re.compile(r"(\d+d(\d+|%)\s?[\+\-x]?\s*\d*|DC\s*\d+|[\+\-]\d+\s+([A-Za-z]+\s+)?(\s*armor\s+)?(bonus|penalty)|[\+\-]?\d+\s+AC|AC\s+\d+|(\d+,)?\d+\s+(cp|sp|gp|pp|rounds?|feet))")
    rangeOnly = re.compile(r".*? \((.*?)\)")
    splitSubschools = re.compile(r"\s+or\s+|,\s*")
    snakeSigilSpecial = re.compile(r"(Abjuration|Conjuration|Enchantment|Evocation|Illusion|Necromancy|Transmutation):\s+(.+?)(?=\s*(Abjuration|Conjuration|Enchantment|Evocation|Illusion|Necromancy|Transmutation|$):?)")
    domains = re.compile(r"(.+?)\s+\((\d+)\),?\ ?")
    for row in rows:
        record = {}
        for key, cell in zip(first_row, row):
            if cell.data_type == 's':
                record[key] = cell.value.strip()
            elif cell.data_type == 'n':
                if cell.value == None:
                    record[key] = ""
                else:
                    record[key] = str(int(float(cell.value)))
            else:
                record[key] = cell.value

        card = {}
        school = record["school"]
        card["school"] = school.capitalize() 
        card["subschools"] = [ x.capitalize() for x in splitSubschools.split(record["subschool"]) ]

        for c in classes:
            if record[c] != "NULL":
                card[c] = record[c]

        if record["alchemist"] != "NULL" and int(card["alchemist"]) > 0 and int(card["alchemist"]) <= 6:
            card["investigator"] = card["alchemist"]

        if record["bard"] != "NULL" and int(card["bard"]) > 0 and int(card["bard"]) <= 6:
            card["skald"] = card["bard"]

        if record["cleric"] != "NULL" and int(card["cleric"]) <= 6:
            card["warpriest"] = card["cleric"]

        card["id"] = int(record["id"])
        card["name"] = titlecase(record["name"])

        card["components_summary"] = details.sub('', record["components"])
        card["components"] = []
        if record["verbal"] != "0":
            card["components"].append("V")
        if record["somatic"] != "0":
            card["components"].append("S")
        if record["material"] != "0":
            card["components"].append("M")
        if record["focus"] != "0":
            card["components"].append("F")
        if record["divine_focus"] != "0":
            card["components"].append("DF")
        card["materials"] = []
        for m in materials.finditer(record["components"]):
            card["materials"].append({ "kind" : m.group(1), "description" : m.group(2).capitalize() })

        card["time"] = shorten_casting_time(record["casting_time"])
        card["tc"] = get_time_category(card["time"])
        card["ts"] = is_time_special(card["time"])
        card["range"] = shorten(rangeOnly.sub(r"\1", record["range"]))
        card["duration"] = shorten(record["duration"])
        card["save"] = shorten(record["saving_throw"])
        card["resistance"] = shorten(record["spell_resistence"])
        spellArea = record["area"]
        spellTargets = record["targets"]
        combineAT = spellArea == spellTargets
        card["area_targets"] = []
        if combineAT == False and spellArea != "" and spellArea != "None":
            card["area_targets"].append({ "kind" : "A", "description" : shorten(spellArea, False) })
        if spellTargets != "" and spellTargets != "None":
            if combineAT:
                card["area_targets"] = [{ "kind" : "A/W", "description" : shorten(spellTargets, False) }]
            else:
                card["area_targets"].append({ "kind" : "W", "description" : shorten(spellTargets, False) })

        record["description"] = dicedc.sub(r"<strong>\1</strong>", record["description"])
        card["description"] = conditions(record["description"])
        card["source"] = record["source"]
        card["sla"] = record["SLA_Level"]

        descript = []
        for d in descriptors:
            if record[d] == "1":
                descript.append(titlecase(d.replace('_', ' ')))
        if len(descript) > 0:
            card["descriptors"] = descript

        domain = {}
        for m in domains.finditer(record["domain"]):
            domain[m.group(1)] = m.group(2)
        if len(domain) > 0:
            card["domain"] = domain

        bloodline = {}
        for m in domains.finditer(record["bloodline"]):
            bloodline[m.group(1)] = int(m.group(2)) // 2
        if len(bloodline) > 0:
            card["bloodline"] = bloodline

        patron = {}
        for m in domains.finditer(record["patron"]):
            patron[m.group(1)] = int(m.group(2)) // 2
        if len(patron) > 0:
            card["patron"] = patron

        if card["name"] == "Lissalan Snake Sigil":
            originalName = card["name"]
            originalDescription = card["description"].replace("There are seven variants of this spell, one for each of the Thassilonian schools of magic. Each", "This spell")
            originalDescription = snakeSigilSpecial.sub("", originalDescription)
            for m in snakeSigilSpecial.finditer(card["description"]):
                school = m.group(1)
                description = m.group(2)
                card["name"] = originalName + ", " + school
                card["school"] = school
                card["description"] = description + " " + originalDescription
                cards[card["id"]] = card.copy()
            continue

        cards[card["id"]] = card.copy()

    print(dumps(cards, ensure_ascii=False))

read_spells('spell_full.xlsx')
